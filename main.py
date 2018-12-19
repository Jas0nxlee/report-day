import openpyxl
wb = openpyxl.load_workbook('in.xlsx')
ws2 = wb.create_sheet()
ws = wb.active

allworkname = ['部门','姓名','总工时']        #代号lists
allworktime = []        #工时list
cloumn_length = len(list(ws.rows))-2      #总表长度


#——————————————————————————————————————————生成总代号list—————allworkname———————————————————————————————————————
for n in range(cloumn_length):
    #计算总工时
    worktime = 0        #初始化总工时计数器
    col = 9     #数据位置指针初始化 
    for y in range(4):      #每月4次工时累加
        for x in range(7):      #每周7次工时累加
            ex1 = ws.cell(row=n+3,column=col)     #项目代号位置
            ex2 = ws.cell(row=n+3,column=col+1)       #项目工时位置
            col = col + 3       #数据位置指针周内移动
            if ex2.value != None:       #排除空单元格
                worktime = worktime + ex2.value    #总工时累加
        col = col + 1       #数据位置指针移动至周开始

    #计算项目工时占比
    worklist = []       #工时list初始化
    col = 9
    for y in range(4):
        for x in range(7):
            ex1 = ws.cell(row=n+3,column=col)
            ex2 = ws.cell(row=n+3,column=col+1)
            col = col + 3
            if ex2.value != None:
                workin = ex2.value/worktime        #工时占比计算
                #print(ex1.value,workin)         #打着玩
                if ex1.value in worklist:      #判断重复项目
                    u = worklist.index(ex1.value)+1    
            
                    worklist[u] = worklist[u] + workin #重复项目工时占比累加
                else:
                    worklist.append(ex1.value)      #写项目代号
                    worklist.append(workin)     #写工时
        col = col + 1
    print(worklist)
    #worklist内项目及工时占比拆分
    pj = []     
    works = []
    for c in range(0,len(worklist),2):
        pj.append(worklist[c])
        works.append(worklist[c+1])

    #汇总项目代号
    for idx in pj:
        if idx not in allworkname:
            allworkname.append(idx)
#print(allworkname)
ws2.append(allworkname)

#————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————

for n in range(cloumn_length):
    #计算总工时
    datarow = []
    worktime = 0        #初始化总工时计数器
    col = 9     #数据位置指针初始化 
    for y in range(4):      #每月4次工时累加
        for x in range(7):      #每周7次工时累加
            ex1 = ws.cell(row=n+3,column=col)     #项目代号位置
            ex2 = ws.cell(row=n+3,column=col+1)       #项目工时位置
            col = col + 3       #数据位置指针周内移动
            if ex2.value != None:       #排除空单元格
                worktime = worktime + ex2.value        #总工时累加
        col = col + 1       #数据位置指针移动至周开始

    datarow.append(ws.cell(row=n+3,column=1).value)
    datarow.append(ws.cell(row=n+3,column=2).value)
    datarow.append(worktime)


    #计算项目工时占比
    worklist = []       #工时list初始化
    col = 9
    for y in range(4):
        for x in range(7):
            ex1 = ws.cell(row=n+3,column=col)
            ex2 = ws.cell(row=n+3,column=col+1)
            col = col + 3
            if ex2.value != None:
                workin = ex2.value/worktime        #工时占比计算
                #print(ex1.value,workin)         #打着玩
                if ex1.value in worklist:      #判断重复项目
                    u = worklist.index(ex1.value)+1    
            
                    worklist[u] = worklist[u] + workin #重复项目工时累加
                else:
                    worklist.append(ex1.value)      #写项目代号
                    worklist.append(workin)     #写工时
        col = col + 1
    
    #worklist内项目及工时占比拆分
    pj = []     
    works = []
    for c in range(0,len(worklist),2):
        pj.append(worklist[c])
        works.append(worklist[c+1])
    #填充datarow list
    for z in allworkname:
        datarow.append(0)
    datarow.pop()
    datarow.pop()
    datarow.pop()
    #筛选数据赋值
    for indx1 in allworkname:
        for indx2 in pj:
            if indx2 == indx1:
                xx = pj.index(indx2)
                yy = allworkname.index(indx1)
                datarow[yy] = works[xx]
                
    
    ws2.append(datarow)
    print(datarow)
wb.save(r'out.xlsx')