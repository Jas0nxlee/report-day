
import openpyxl
wb = openpyxl.load_workbook('in.xlsx')
ws2 = wb.create_sheet()
ws = wb.active
workname = []
worknameout = []
worktime = 0
null = ''
cloumn_length = len(list(ws.rows))-2      #总表长度

#——————————————————————————————————             总工时———————————————————————————————————————
for n in range(cloumn_length):    #计算总工时
    col = 10
    for x in range(5):      #每周7次工时累加
        ex1 = ws.cell(row=n+3,column=col+1)     #项目代号位置
        ex2 = ws.cell(row=n+3,column=col+2)       #项目工时位置
        col = col + 3
        if ex2.value != null and ex2.value != None:       #排除空单元格
            worktime += round(float(ex2.value),2)   #总工时累加
            if ex1.value not in workname:
                workname.append(ex1.value)
worknameout.append("姓名")
worknameout.append("部门")
worknameout.append("总工时")
worknameout.extend(workname)
#print(worknameout)
ws2.append(worknameout)
#print(worktime)

#—————————————————————————————————————个人项目工时列表————————————————————————————————————————
all = []
aa = []
bbtmp = []
allworktime = 0
for n in range(cloumn_length):
    ex1 = ws.cell(row=n+3,column=1)
    ex2 = ws.cell(row=n+3,column=2)
    if ex1.value not in bbtmp:
        aa.append(ex1.value)
        aa.append(ex2.value)
        bbtmp.append(ex1.value)
        for p in range(cloumn_length):
            ex3 = ws.cell(row=p+3,column=1)
            ex4 = ws.cell(row=p+3,column=2)
            if ex3.value == ex1.value:
                col = 10
                for x in range(5):
                    ex5 = ws.cell(row=p+4,column=col+1)
                    ex6 = ws.cell(row=p+4,column=col+2)
                    col = col + 3
                    if ex6.value != null and ex6.value != None:
                        allworktime += round(float(ex6.value),2)
                        aa.append(ex5.value)
                        aa.append(ex6.value)

        aa.insert(2,allworktime)
        all.append(aa)
        aa = []
        allworktime = 0


#—————————————————————————————————单人工时占比—————————————————————————————————

for n in range(len(all)):
    pj = []
    works = []
    aaa = all[n]
    datarowout = []
    datarowout.append(aaa[0])
    datarowout.append(aaa[1])
    datarowout.append(aaa[2])
    datarow = []
  #  datarow.append(aaa[0])
  #  datarow.append(aaa[1])
  #  datarow.append(aaa[2])
    for m in range(3,len(aaa),2):
        aaa[m+1]= float(aaa[m+1])/aaa[2]
        if aaa[m] in datarow:
            u = datarow.index(aaa[m])+1
            datarow[u] += aaa[m+1]
        else:
            datarow.append(aaa[m])
            datarow.append(aaa[m+1])

    for q in range(0,len(datarow),2):
        pj.append(datarow[q])
        works.append(datarow[q+1])

    for r in range(len(workname)):
        datarowout.append(0)

    for indx1 in worknameout:
       for indx2 in pj:
            if indx2 == indx1:
               xx = pj.index(indx2)
               yy = worknameout.index(indx2)
               datarowout[yy] = works[xx]
#    print(datarowout)
    ws2.append(datarowout)
wb.save(r'out.xlsx')


